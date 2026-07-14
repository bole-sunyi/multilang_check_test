from __future__ import annotations

"""
把点击后截图落到 Excel 模板里的工具模块。

这个文件专门解决两件事：
1. 按点击顺序把截图名称写进 A 列；
2. 把对应截图按比例缩放后插入 B 列。

这样业务同学打开模板时，不需要理解 Airtest 日志结构，
只看 Excel 就能按顺序核对每一步点击后的页面状态。
"""

from collections.abc import Mapping, Sequence
import os
from pathlib import Path
import shutil
from typing import TYPE_CHECKING, Any, Protocol, TypedDict, cast

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from .screenshot_utils import build_resized_png_buffer
from .paths import get_run_artifacts_dir


class ClickSnapshotRecord(TypedDict, total=False):
    module_name: str
    step_index: int
    step_name: str
    action: str
    image_name: str
    image_path: str
    locate_method: str


class _AnchorMarkerLike(Protocol):
    row: int
    col: int

# 这些常量统一管理模板位置、工作表名和图片尺寸。
# 以后如果模板地址或缩略图大小要调整，优先改这里即可。
LEGACY_EXCEL_TEMPLATE_PATH = Path("/Users/sunyi/Downloads/多语测试模板.xlsx")
EXCEL_TEMPLATE_PATH_ENV_NAMES = ("MULTILANG_EXCEL_TEMPLATE_PATH", "EXCEL_TEMPLATE_PATH")
LOCAL_EXCEL_OUTPUT_NAME = "多语测试模板.xlsx"
DEFAULT_SHEET_NAME = "Sheet1"
DEFAULT_START_ROW = 3
NAME_COLUMN = "A"
IMAGE_COLUMN = "B"
LOCATOR_COLUMN = "C"
MAX_IMAGE_WIDTH_PX = 520
MAX_IMAGE_HEIGHT_PX = 390
IMAGE_CELL_PADDING_PX = 8
MIN_NAME_COLUMN_WIDTH = 24
MIN_IMAGE_COLUMN_WIDTH = 75
MIN_LOCATOR_COLUMN_WIDTH = 18


def resolve_excel_template_path(workbook_path: Path | None = None) -> Path:
    """
    决定本次截图要写入哪个 Excel 文件。

    优先级：
    1. 调用方显式传入的 workbook_path；
    2. 环境变量 MULTILANG_EXCEL_TEMPLATE_PATH / EXCEL_TEMPLATE_PATH；
    3. 新产物目录下固定的 `多语测试模板.xlsx`。
    """
    if workbook_path is not None:
        return workbook_path.expanduser().resolve()

    for env_name in EXCEL_TEMPLATE_PATH_ENV_NAMES:
        env_path = os.environ.get(env_name, "").strip()
        if env_path:
            return Path(env_path).expanduser().resolve()

    return _resolve_default_workbook_path()


def _resolve_default_workbook_path() -> Path:
    """默认把 Excel 输出到产物目录固定模板文件。"""
    output_path = get_run_artifacts_dir() / LOCAL_EXCEL_OUTPUT_NAME
    if not output_path.exists() and LEGACY_EXCEL_TEMPLATE_PATH.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(LEGACY_EXCEL_TEMPLATE_PATH, output_path)
    return output_path.resolve()


def export_click_screenshots_to_excel(
    click_snapshot_records: Sequence[ClickSnapshotRecord | Mapping[str, object]],
    workbook_path: Path | None = None,
    sheet_name: str = DEFAULT_SHEET_NAME,
    start_row: int = DEFAULT_START_ROW,
) -> Path:
    """兼容旧调用名；现在实际按模块 sheet 写入截图记录。"""
    return export_module_screenshots_to_excel(
        click_snapshot_records,
        workbook_path=workbook_path,
        sheet_name=sheet_name,
        start_row=start_row,
    )


def export_module_screenshots_to_excel(
    snapshot_records: Sequence[ClickSnapshotRecord | Mapping[str, object]],
    workbook_path: Path | None = None,
    sheet_name: str = DEFAULT_SHEET_NAME,
    start_row: int = DEFAULT_START_ROW,
) -> Path:
    """
    把单模块截图按顺序写入 Excel 的指定 sheet。

    约定：
    1. 从 A 列写步骤/截图名称；
    2. 从 B 列插入对应截图；
    3. 从 C 列写定位方式，方便确认本次步骤是否通过 Poco 定位；
    4. 默认从第 3 行开始，保留模板已有表头。
    """
    workbook_path = resolve_excel_template_path(workbook_path)

    from openpyxl import Workbook, load_workbook

    if workbook_path.exists():
        workbook = load_workbook(workbook_path)
    else:
        # 多语言检查通常临时跑脚本验证截图，模板缺失时直接创建一份最小可用工作簿，
        # 避免一次截图流程因为外部 Excel 文件不存在而中断。
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.active.title = sheet_name

    worksheet = _resolve_worksheet(workbook, sheet_name)
    _ensure_minimal_headers(worksheet, start_row=start_row)

    # 先清空旧内容，再写本轮结果，避免多次执行后旧图和新图叠在一起。
    _clear_output_area(worksheet, start_row=start_row)
    max_name_width = MIN_NAME_COLUMN_WIDTH
    max_image_width = MIN_IMAGE_COLUMN_WIDTH
    max_locator_width = MIN_LOCATOR_COLUMN_WIDTH

    for offset, record in enumerate(snapshot_records):
        row = start_row + offset
        image_name = _build_display_name(record)
        image_path = Path(_read_record_text(record, "image_path")).expanduser().resolve()
        # 某条记录如果名称为空，或者图片文件已经不存在，就跳过。
        # 这样整批导出不会因为一张异常图片而全部失败。
        if not image_name or not image_path.exists():
            continue

        worksheet[f"{NAME_COLUMN}{row}"] = image_name
        inserted_width, row_height = _insert_image(worksheet, f"{IMAGE_COLUMN}{row}", image_path)
        locator_text = _build_locator_text(record)
        worksheet[f"{LOCATOR_COLUMN}{row}"] = locator_text
        worksheet.row_dimensions[row].height = row_height
        max_name_width = max(max_name_width, min(len(image_name) * 1.2, 60))
        max_image_width = max(max_image_width, _pixels_to_excel_width(inserted_width))
        max_locator_width = max(max_locator_width, min(len(locator_text) * 1.2, 45))

    worksheet.column_dimensions[NAME_COLUMN].width = max_name_width
    worksheet.column_dimensions[IMAGE_COLUMN].width = max_image_width
    worksheet.column_dimensions[LOCATOR_COLUMN].width = max_locator_width

    workbook.save(workbook_path)
    return workbook_path


def _resolve_worksheet(workbook: Any, sheet_name: str) -> "Worksheet":
    """获取或创建模块 sheet。"""
    workbook_obj = cast(object, workbook)
    sheetnames = cast(list[str], getattr(workbook_obj, "sheetnames", []))
    if sheet_name in sheetnames:
        return cast("Worksheet", workbook[sheet_name])
    if not sheetnames:
        return cast("Worksheet", workbook.create_sheet(sheet_name))
    active = cast("Worksheet", getattr(workbook_obj, "active"))
    if active.title == DEFAULT_SHEET_NAME and not active.max_row:
        active.title = sheet_name
        return active
    return cast("Worksheet", workbook.create_sheet(sheet_name))


def _ensure_minimal_headers(worksheet: "Worksheet", start_row: int) -> None:
    """模板缺少表头时补一份最小表头，便于直接交给多语言检查使用。"""
    header_row = max(start_row - 1, 1)
    if not worksheet[f"{NAME_COLUMN}{header_row}"].value:
        worksheet[f"{NAME_COLUMN}{header_row}"] = "截图名称"
    if not worksheet[f"{IMAGE_COLUMN}{header_row}"].value:
        worksheet[f"{IMAGE_COLUMN}{header_row}"] = "截图"
    if not worksheet[f"{LOCATOR_COLUMN}{header_row}"].value:
        worksheet[f"{LOCATOR_COLUMN}{header_row}"] = "定位方式"


def _clear_output_area(worksheet: "Worksheet", start_row: int) -> None:
    """清空输出区域中的旧文件名、旧图片和旧行高。"""
    max_row = max(worksheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        worksheet[f"{NAME_COLUMN}{row}"] = None
        worksheet[f"{IMAGE_COLUMN}{row}"] = None
        worksheet[f"{LOCATOR_COLUMN}{row}"] = None
        worksheet.row_dimensions[row].height = None

    remained_images: list[object] = []
    for image in _get_worksheet_images(worksheet):
        anchor = cast(object | None, getattr(image, "anchor", None))
        marker = cast(_AnchorMarkerLike | None, getattr(anchor, "_from", None))
        if marker is None:
            remained_images.append(image)
            continue
        row_index = int(marker.row) + 1
        col_index = int(marker.col) + 1
        # A/B/C 列、起始行之后的图片都属于本工具的输出区域，重新导出时要移除。
        if row_index >= start_row and col_index in (1, 2, 3):
            continue
        remained_images.append(image)
    setattr(worksheet, "_images", remained_images)


def _insert_image(worksheet: "Worksheet", cell: str, image_path: Path) -> tuple[int, float]:
    """
    把图片按比例缩放后插入指定单元格。

    返回值是 `(缩放后宽度像素, 建议行高 points)`。
    """
    from openpyxl.drawing.image import Image as OpenpyxlImage

    max_width_px, max_height_px = _get_image_cell_limits(worksheet, cell)
    buffer, width, height = build_resized_png_buffer(
        image_path,
        max_width_px=max_width_px,
        max_height_px=max_height_px,
        prefer_landscape=True,
    )
    excel_image = OpenpyxlImage(buffer)
    excel_image.width = width
    excel_image.height = height
    worksheet.add_image(excel_image, cell)

    # openpyxl 行高单位是 point，近似按 0.75 换算像素。
    return width + IMAGE_CELL_PADDING_PX, max((height + IMAGE_CELL_PADDING_PX) * 0.75, 20)


def _get_image_cell_limits(worksheet: "Worksheet", cell: str) -> tuple[int, int]:
    """返回本工具期望的截图单元格尺寸，不继承旧表格里可能过小的列宽/行高。"""
    _ = (worksheet, cell)
    max_width_px = max(MAX_IMAGE_WIDTH_PX - IMAGE_CELL_PADDING_PX, 1)
    max_height_px = max(MAX_IMAGE_HEIGHT_PX - IMAGE_CELL_PADDING_PX, 1)
    return max_width_px, max_height_px


def _get_worksheet_images(worksheet: "Worksheet") -> list[object]:
    """安全读取 worksheet 当前挂载的图片列表。"""
    images = cast(object, getattr(worksheet, "_images", []))
    return list(cast(list[object], images)) if isinstance(images, list) else list(cast(list[object], images or []))


def _read_record_text(record: ClickSnapshotRecord | Mapping[str, object], key: str) -> str:
    """把截图记录里的值安全转成字符串，避免类型检查器把动态字段标红。"""
    value = record.get(key, "")
    return value.strip() if isinstance(value, str) else ""


def _build_display_name(record: ClickSnapshotRecord | Mapping[str, object]) -> str:
    step_name = _read_record_text(record, "step_name")
    image_name = _read_record_text(record, "image_name")
    # Excel A 列直接展示 YAML 中的 name。
    # 这样测试同学看到的截图名称、atw_screen 文件名、报告里的截图名称能保持一致。
    return step_name or image_name


def _build_locator_text(record: ClickSnapshotRecord | Mapping[str, object]) -> str:
    return _read_record_text(record, "locate_method") or "unknown"


def _pixels_to_excel_width(width_px: int) -> float:
    """
    把像素宽度近似换算成 Excel 列宽单位。

    这里采用经验公式，目标不是像素级精确，而是让横版图能完整显示在单元格里。
    """
    return max(round((max(width_px, 12) + 5) / 7, 2), MIN_IMAGE_COLUMN_WIDTH)
